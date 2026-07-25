#!/usr/bin/env python3
"""
Excel-to-CSV Database Synchronizer for Soul Blueprint Matrix.
Author: Ahmad Hassan (B-Ted)
"""

import os
import sys
import csv
import shutil
import re
from datetime import datetime, date

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
BACKUP_DIR = os.path.join(SCRIPT_DIR, "backups")


def read_xlsx_fallback(filepath):
    """Fallback XLSX reader using standard library zipfile and xml.etree."""
    import zipfile
    import xml.etree.ElementTree as ET

    sheets = {}
    with zipfile.ZipFile(filepath, 'r') as z:
        shared_strings = []
        if 'xl/sharedStrings.xml' in z.namelist():
            tree = ET.fromstring(z.read('xl/sharedStrings.xml'))
            for elem in tree.iter():
                if elem.tag.endswith('t') and elem.text is not None:
                    shared_strings.append(elem.text)

        sheet_names = {}
        if 'xl/workbook.xml' in z.namelist():
            wb_tree = ET.fromstring(z.read('xl/workbook.xml'))
            ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            for sheet in wb_tree.findall('.//s:sheet', ns):
                s_name = sheet.attrib.get('name')
                r_id = sheet.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                if s_name and r_id:
                    sheet_names[r_id] = s_name

        ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        sheet_files = [f for f in z.namelist() if f.startswith('xl/worksheets/sheet') and f.endswith('.xml')]
        sheet_files.sort()

        for idx, fpath in enumerate(sheet_files):
            r_id = f"rId{idx+1}"
            title = sheet_names.get(r_id, f"Sheet{idx+1}")
            sheet_tree = ET.fromstring(z.read(fpath))
            rows = []
            for row_elem in sheet_tree.findall('.//s:row', ns):
                row_vals = []
                for c_elem in row_elem.findall('s:c', ns):
                    t_attr = c_elem.attrib.get('t')
                    v_elem = c_elem.find('s:v', ns)
                    val = v_elem.text if v_elem is not None else None
                    if t_attr == 's' and val is not None and val.isdigit():
                        s_idx = int(val)
                        val = shared_strings[s_idx] if s_idx < len(shared_strings) else val
                    elif t_attr == 'inlineStr':
                        is_elem = c_elem.find('s:is', ns)
                        val = is_elem.find('s:t', ns).text if is_elem is not None else val
                    row_vals.append(val)
                rows.append(row_vals)
            sheets[title] = rows

    return sheets


def get_excel_file(data_dir):
    search_paths = [
        os.path.join(data_dir, "interpretations.xlsx"),
        os.path.join(data_dir, "Interpretations.xlsx"),
        os.path.join(data_dir, "Interpretations backup.xlsx"),
        os.path.join(data_dir, "customer_s-editited", "Interpretations backup.xlsx"),
    ]
    
    for path in search_paths:
        if os.path.isfile(path):
            return path
            
    if os.path.isdir(data_dir):
        xlsx_files = [os.path.join(data_dir, f) for f in os.listdir(data_dir) if f.endswith(".xlsx")]
        if xlsx_files:
            xlsx_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
            return xlsx_files[0]
        
    return None


def create_backup(csv_path, backup_dir):
    if not os.path.isfile(csv_path):
        return None
        
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"interpretations_backup_{timestamp}.csv")
    
    try:
        shutil.copy2(csv_path, backup_path)
        return backup_path
    except Exception as e:
        print(f"[WARNING] Could not create backup file: {e}")
        return None


def map_excel_to_csv_columns(pos, pos_meaning, excel_section, is_compat=False):
    pos = str(pos).strip()
    excel_section = str(excel_section).strip()
    
    pos_upper = pos.upper().replace(' ', '')
    is_age = pos_upper.startswith('AGE')
    
    if pos_upper == 'FORECAST':
        sec_lower = excel_section.lower().strip()
        section = 'theme'
        if 'watch' in sec_lower or 'trap' in sec_lower or 'risk' in sec_lower:
            section = 'watch_out'
        elif 'recommendation' in sec_lower or 'rec' in sec_lower or 'unlock' in sec_lower:
            section = 'recommendations'
        return 'FORECAST', 'forecast', section
        
    if pos_upper == 'COMPAT_FORECAST':
        sec_lower = excel_section.lower().strip()
        section = 'theme'
        if 'watch' in sec_lower or 'trap' in sec_lower or 'risk' in sec_lower:
            section = 'watch_out'
        elif 'recommendation' in sec_lower or 'rec' in sec_lower or 'unlock' in sec_lower:
            section = 'recommendations'
        return 'COMPAT_FORECAST', 'compat_forecast', section

    if is_age:
        age_num_str = pos_upper[3:].replace(',', '.')
        pos_clean = f"age{age_num_str}"
        module = 'forecast'
        sec_lower = excel_section.lower().strip()
        if sec_lower in ['interpretation', '']:
            section = 'yearly'
        else:
            section = sec_lower.replace(' ', '_').strip('_')
            if not section:
                section = 'yearly'
        if is_compat:
            module = f"compat_{module}"
        return pos_clean, module, section

    pos_upper = pos.upper()
    is_program = '-' in pos or ',' in pos or pos_upper == 'PROGRAM' or 'PROGRAM' in str(pos_meaning).upper()
    if is_program:
        module = 'compat_programs' if is_compat else 'programs'
        section = excel_section.lower().replace(' ', '_').strip('_')
        if not section:
            section = 'program_meaning'
        # Clean position key for programs e.g. "L2 - L1 - L" -> "L2-L1-L"
        clean_parts = [p.strip().upper() for p in re.split(r'[\s,-]+', pos) if p.strip()]
        pos_clean = '-'.join(clean_parts) if clean_parts else pos_upper
        return pos_clean, module, section

    sec_lower = excel_section.lower().strip()

    if is_compat:
        if 'general' in sec_lower or sec_lower in ['interpretation', '']:
            return pos_upper, 'compat_general', 'meaning'
        elif 'love' in sec_lower or 'relationship' in sec_lower:
            return pos_upper, 'compat_love', 'meaning'
        elif 'karma' in sec_lower:
            return pos_upper, 'compat_karma', 'meaning'
        elif 'finance' in sec_lower or 'money' in sec_lower or 'wealth' in sec_lower:
            return pos_upper, 'compat_finance', 'meaning'
        elif 'forecast' in sec_lower or 'yearly' in sec_lower:
            return pos_upper, 'compat_forecast', 'yearly'
        else:
            module = 'compat_' + sec_lower.replace(' ', '_').strip('_')
            if not module:
                module = 'compat_general'
            return pos_upper, module, 'meaning'

    # Single DOB mapping
    if sec_lower in ['interpretation', '']:
        default_map = {
            'B': ('core', 'meaning'),
            'C': ('core', 'meaning'),
            'D': ('core', 'meaning'),
            'E': ('core', 'meaning'),
            'F': ('core', 'meaning'),
            'F2': ('core', 'meaning'),
            'G': ('core', 'meaning'),
            'G2': ('core', 'meaning'),
            'H': ('core', 'meaning'),
            'H1': ('core', 'meaning'),
            'I': ('core', 'meaning'),
            'I1': ('core', 'meaning'),
            'J': ('core', 'meaning'),
            'K': ('purpose', 'gifts'),
            'L': ('relationships', 'lesson'),
            'M': ('relationships', 'partner'),
            'N': ('karma', 'karmic'),
            'O': ('core', 'meaning'),
            'P': ('core', 'meaning'),
            'Q': ('core', 'meaning'),
            'R': ('relationships', 'relationship_problems'),
            'R1': ('relationships', 'nature_of_the_relationship'),
            'R2': ('money', 'activation'),
            'S': ('relationships', 'wound'),
            'T': ('relationships', 'meaning'),
        }
        if pos_upper in default_map:
            module, section = default_map[pos_upper]
        else:
            module, section = 'core', 'meaning'
        return pos_upper, module, section

    clean_sec = re.sub(r'[^a-z0-9\s]', ' ', sec_lower)
    clean_sec = ' '.join(clean_sec.split())
    
    if pos_upper in ['R1', 'L', 'M', 'S']:
        module = 'relationships'
    elif pos_upper in ['R2']:
        module = 'money'
    elif pos_upper == 'N':
        module = 'karma'
    elif 'relationship' in clean_sec or 'love' in clean_sec:
        module = 'relationships'
    elif 'karma' in clean_sec:
        module = 'karma'
    elif 'money' in clean_sec or 'finance' in clean_sec or 'wealth' in clean_sec:
        module = 'money'
    elif 'purpose' in clean_sec:
        module = 'purpose'
    elif 'forecast' in clean_sec:
        module = 'forecast'
    else:
        module = 'relationships' if pos_upper == 'R' else 'core'
        
    if any(w in clean_sec for w in ['problem', 'problems', 'wound', 'wounds', 'crisis', 'block', 'blocks', 'challenge', 'shadow']):
        if module == 'relationships':
            section = 'wound'
        elif module == 'money':
            section = 'block'
        else:
            section = 'shadow'
    elif any(w in clean_sec for w in ['partner', 'partners', 'attract', 'attraction', 'dynamics']):
        section = 'partner'
    elif any(w in clean_sec for w in ['lesson', 'lessons']):
        section = 'lesson'
    elif any(w in clean_sec for w in ['money_flow', 'flow', 'income']):
        section = 'money_flow'
    elif any(w in clean_sec for w in ['activation', 'activate']):
        section = 'activation'
    elif any(w in clean_sec for w in ['positive', 'gift', 'gifts']):
        section = 'positive' if module == 'core' else 'gifts'
    elif any(w in clean_sec for w in ['meaning', 'general', 'interpretation', 'description']):
        section = 'meaning'
    else:
        section_part = clean_sec
        for kw in ['relationships', 'relationship', 'love', 'money', 'finance', 'karma', 'core', 'purpose', 'forecast']:
            section_part = section_part.replace(kw, '').strip()
        section = section_part.replace(' ', '_').strip('_')
        if not section:
            section = 'meaning' if module != 'relationships' else 'partner'
            
    return pos_upper, module, section


def main():
    print("=" * 60)
    print("      Soul Blueprint Matrix - Excel Database Sync Tool")
    print("=" * 60)
    
    excel_file = get_excel_file(DATA_DIR)
    if not excel_file:
        print("[ERROR] Could not find any interpretations spreadsheet!")
        print(f"Please place your Excel file (e.g. interpretations.xlsx) in the '{DATA_DIR}' folder.")
        input("\nPress Enter to exit...")
        sys.exit(1)
        
    print(f"[INFO] Found Excel database: {excel_file}")
    
    loaded_sheets = {}
    if HAS_OPENPYXL:
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                loaded_sheets[sname] = list(ws.iter_rows(values_only=True))
            wb.close()
        except PermissionError:
            print("\n[ERROR] Permission denied opening the Excel file.")
            print("This usually happens if the file is currently open in Microsoft Excel.")
            print("Please CLOSE the Excel file and run this tool again.")
            input("\nPress Enter to exit...")
            sys.exit(1)
        except Exception as e:
            print(f"[INFO] openpyxl load failed ({e}), falling back to built-in XLSX reader...")
            loaded_sheets = read_xlsx_fallback(excel_file)
    else:
        print("[INFO] Using built-in Python XLSX reader engine...")
        try:
            loaded_sheets = read_xlsx_fallback(excel_file)
        except Exception as e:
            print(f"\n[ERROR] Failed to read Excel database: {e}")
            input("\nPress Enter to exit...")
            sys.exit(1)

    sheet_names = list(loaded_sheets.keys())
    sheets_to_process = []
    single_sheet = "Master_Database" if "Master_Database" in sheet_names else sheet_names[0]
    sheets_to_process.append((single_sheet, False))
    
    compat_sheet = None
    if "Compatibility" in sheet_names:
        compat_sheet = "Compatibility"
    else:
        for s in sheet_names:
            if s.lower() != single_sheet.lower() and "compat" in s.lower():
                compat_sheet = s
                break
                
    if compat_sheet:
        sheets_to_process.append((compat_sheet, True))
        print(f"[INFO] Found compatibility sheet: '{compat_sheet}'")
    else:
        print("[INFO] No compatibility sheet found. Processing single DOB sheet only.")
        
    required_cols = {
        'position': ['position', 'pos', 'node'],
        'position meaning': ['position meaning', 'meaning', 'description'],
        'section': ['section', 'category', 'tab'],
        'number': ['number', 'val', 'num', 'key'],
        'interpretation text': ['interpretation text', 'text', 'content', 'interpretation']
    }
    
    csv_filename = os.path.join(DATA_DIR, "interpretations.csv")
    backup_path = create_backup(csv_filename, BACKUP_DIR)
    if backup_path:
        print(f"[OK] Pre-existing CSV backed up to: {backup_path}")
        
    mapped_count = 0
    empty_count = 0
    invalid_rows = []
    active_positions = set()
    
    try:
        with open(csv_filename, 'w', encoding='utf-8', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['position', 'module', 'section', 'number', 'text'])
            
            for sheet_name, is_compat in sheets_to_process:
                rows = loaded_sheets.get(sheet_name, [])
                print(f"[INFO] Reading sheet: '{sheet_name}'")
                
                if not rows:
                    print(f"[WARNING] Sheet '{sheet_name}' is empty.")
                    continue
                    
                headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
                
                col_indices = {}
                for col_key, aliases in required_cols.items():
                    found_idx = None
                    for alias in aliases:
                        if alias in headers:
                            found_idx = headers.index(alias)
                            break
                    if found_idx is None:
                        for i, h in enumerate(headers):
                            if any(alias in h for alias in aliases):
                                found_idx = i
                                break
                    if found_idx is None:
                        print(f"[ERROR] Could not find the '{col_key}' column in Excel sheet '{sheet_name}'!")
                        print(f"Spreadsheet columns must include headers like: Position, Section, Number, Interpretation Text.")
                        print(f"Current columns found: {headers}")
                        input("\nPress Enter to exit...")
                        sys.exit(1)
                    col_indices[col_key] = found_idx
                    
                pos_idx = col_indices['position']
                meaning_idx = col_indices['position meaning']
                sec_idx = col_indices['section']
                num_idx = col_indices['number']
                text_idx = col_indices['interpretation text']
                
                for row_num, row in enumerate(rows[1:], start=2):
                    p_val = row[pos_idx] if pos_idx < len(row) else None
                    p_mean = row[meaning_idx] if meaning_idx < len(row) else None
                    sec_val = row[sec_idx] if sec_idx < len(row) else None
                    num_val = row[num_idx] if num_idx < len(row) else None
                    text_val = row[text_idx] if text_idx < len(row) else None
                    
                    if p_val is None or str(p_val).strip() == "":
                        continue
                    if text_val is None or str(text_val).strip() == "":
                        empty_count += 1
                        continue
                        
                    p_val_str = str(p_val).strip() if p_val is not None else ""
                    p_mean_str = str(p_mean).strip() if p_mean is not None else ""
                    is_age = p_val_str.upper().replace(' ', '').startswith('AGE')
                    is_program = (not is_age) and ('-' in p_val_str or ',' in p_val_str or p_val_str.upper() == 'PROGRAM' or 'PROGRAM' in p_mean_str.upper())
                    
                    if is_program or is_age:
                        if is_age:
                            try:
                                num = int(float(num_val))
                            except (ValueError, TypeError):
                                num = str(num_val).strip() if num_val is not None else ""
                        else:
                            if isinstance(num_val, (datetime, date)):
                                num_raw = f"{num_val.day}-{num_val.month}-{num_val.year % 100}"
                            else:
                                num_raw = str(num_val).strip() if num_val is not None else ""
                            
                            # Normalize program numbers (e.g., "05-08-21" -> "5-8-21")
                            parts = re.split(r'[\s,-]+', num_raw)
                            norm_parts = []
                            for p in parts:
                                try:
                                    norm_parts.append(str(int(p)))
                                except ValueError:
                                    if p: norm_parts.append(p)
                            num = '-'.join(norm_parts) if norm_parts else num_raw
                    else:
                        try:
                            num = int(float(num_val))
                        except (ValueError, TypeError):
                            invalid_rows.append((sheet_name, row_num, num_val, p_val))
                            continue
                        
                    pos, module, section = map_excel_to_csv_columns(p_val, p_mean, sec_val, is_compat=is_compat)
                    text_clean = str(text_val).strip()
                    
                    writer.writerow([pos, module, section, num, text_clean])
                    mapped_count += 1
                    active_positions.add(pos)
                    
    except PermissionError:
        print(f"\n[ERROR] Permission denied: Could not write to '{csv_filename}'.")
        print("Please close any applications (like text editors or Excel) that might have the file locked.")
        input("\nPress Enter to exit...")
        sys.exit(1)
    except Exception as e:
        print(f"\n[ERROR] Failed to write database: {e}")
        input("\nPress Enter to exit...")
        sys.exit(1)
        
    print("\n" + "=" * 60)
    print("                      SUCCESSFUL SYNC")
    print("=" * 60)
    print(f"[OK] Database updated: {csv_filename}")
    print(f"[OK] Total rows imported: {mapped_count}")
    print(f"[OK] Empty rows skipped: {empty_count}")
    
    if active_positions:
        print(f"[OK] Positions synced: {', '.join(sorted(list(active_positions)))}")
        
    if invalid_rows:
        print(f"\n[WARNING] Skipped {len(invalid_rows)} rows due to invalid 'Number' values:")
        for s_name, r_num, val, pos in invalid_rows[:5]:
            print(f"  - Sheet '{s_name}', Row {r_num}: Number='{val}' (Position {pos})")
        if len(invalid_rows) > 5:
            print(f"  - ... and {len(invalid_rows) - 5} more rows.")
            
    print("\nWorkflow ready. You can now reload your Soul Blueprint Matrix in Chrome.")
    print("=" * 60)
    
    if len(sys.argv) > 1 and sys.argv[1] == '--batch':
        pass


if __name__ == "__main__":
    main()
