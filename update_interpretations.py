#!/usr/bin/env python3
"""
Excel-to-CSV Database Synchronizer for Soul Blueprint Matrix.
Author: Antigravity

This script automates the process of converting the customer-edited Excel spreadsheet
(.xlsx) into the exact CSV structure (position,module,section,number,text) required
by the Soul Blueprint Matrix HTML chart.

Resilience features included:
- Automatic backup of any existing interpretations.csv into a backups/ folder.
- Search fallbacks for multiple Excel filenames/locations.
- Dynamic case-insensitive header column mapping.
- Clean parsing and fallback rules for custom or default sections.
- Graceful error handling for file locks (e.g. if Excel is open), missing dependencies, or bad formats.
"""

import os
import sys
import csv
import shutil
from datetime import datetime

# Required third-party libraries check
try:
    import openpyxl
except ImportError:
    print("=" * 60)
    print("[ERROR] The required library 'openpyxl' is not installed.")
    print("Please run the following command to install it and try again:")
    print("    pip install openpyxl")
    print("=" * 60)
    input("\nPress Enter to exit...")
    sys.exit(1)


def get_excel_file():
    """Locate the Excel file to read, checking fallback options in order."""
    search_paths = [
        "interpretations.xlsx",
        "Interpretations.xlsx",
        "Interpretations backup.xlsx",
        os.path.join("customer_s-editited", "Interpretations backup.xlsx"),
        "customer_s-editited/Interpretations backup.xlsx"
    ]
    
    # Try preferred paths
    for path in search_paths:
        if os.path.isfile(path):
            return path
            
    # Fallback: scan current directory for any .xlsx file
    xlsx_files = [f for f in os.listdir(".") if f.endswith(".xlsx")]
    if xlsx_files:
        # Sort to find the most recently modified one if there are multiple
        xlsx_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        return xlsx_files[0]
        
    return None


def create_backup(csv_path):
    """Back up existing CSV file if it exists."""
    if not os.path.isfile(csv_path):
        return None
        
    backup_dir = "backups"
    os.makedirs(backup_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"interpretations_backup_{timestamp}.csv"
    backup_path = os.path.join(backup_dir, backup_name)
    
    try:
        shutil.copy2(csv_path, backup_path)
        return backup_path
    except Exception as e:
        print(f"[WARNING] Could not create backup file: {e}")
        return None


def map_excel_to_csv_columns(pos, pos_meaning, excel_section):
    """
    Map the human-friendly Excel Position, Position Meaning, and Section columns
    to the specific lowercase module and section names required by the HTML page.
    """
    pos = str(pos).strip()
    excel_section = str(excel_section).strip()
    
    pos_upper = pos.upper()
    is_program = '-' in pos or ',' in pos or pos_upper == 'PROGRAM'
    if is_program:
        module = 'programs'
        section = excel_section.lower().replace(' ', '_').strip('_')
        if not section:
            section = 'program_meaning'
        return pos.upper(), module, section
        
    pos = pos_upper
    
    # 1. Handle "Interpretation" default mapping for general positions
    if excel_section.lower() == 'interpretation':
        default_map = {
            'B': ('core', 'meaning'),
            'C': ('core', 'meaning'),
            'D': ('core', 'meaning'),
            'E': ('core', 'meaning'),
            'F': ('core', 'meaning'),
            'F2': ('core', 'meaning'),
            'G': ('core', 'meaning'),
            'G2': ('core', 'meaning'),
            'H': ('core', 'meaning'),
            'H1': ('core', 'meaning'),
            'I': ('core', 'meaning'),
            'I1': ('core', 'meaning'),
            'J': ('core', 'meaning'),
            'K': ('purpose', 'gifts'),
            'L': ('relationships', 'lesson'),
            'M': ('relationships', 'partner'),
            'N': ('karma', 'karmic'),
            'O': ('core', 'meaning'),
            'P': ('core', 'meaning'),
            'Q': ('core', 'meaning'),
            'R': ('money', 'money_flow'),
            'R1': ('relationships', 'partner'),
            'R2': ('money', 'activation'),
        }
        if pos in default_map:
            return pos, default_map[pos][0], default_map[pos][1]
        else:
            return pos, 'core', 'meaning'
            
    # 2. Parse custom section strings like "Core Meaning", "Core Positive", etc.
    sec_lower = excel_section.lower()
    module = 'core'
    section_part = sec_lower
    
    # Detect the module component of the section header
    if 'core' in sec_lower:
        module = 'core'
        section_part = sec_lower.replace('core', '')
    elif 'compatibility' in sec_lower or 'couples' in sec_lower or 'couple' in sec_lower:
        module = 'compatibility'
        section_part = sec_lower.replace('compatibility', '').replace('couples', '').replace('couple', '')
    elif 'relationship' in sec_lower or 'love' in sec_lower:
        module = 'relationships'
        section_part = sec_lower.replace('relationships', '').replace('relationship', '').replace('love', '')
    elif 'karma' in sec_lower:
        module = 'karma'
        section_part = sec_lower.replace('karma', '')
    elif 'money' in sec_lower:
        module = 'money'
        section_part = sec_lower.replace('money', '')
    elif 'purpose' in sec_lower:
        module = 'purpose'
        section_part = sec_lower.replace('purpose', '')
    elif 'forecast' in sec_lower:
        module = 'forecast'
        section_part = sec_lower.replace('forecast', '')
        
    section_part = section_part.strip()
    
    # Exact mappings for known section words
    section_map = {
        'meaning': 'meaning',
        'positive': 'positive',
        'shadow': 'shadow',
        'healing': 'healing',
        'shadow lessons': 'shadow_lessons',
        'shadow_lessons': 'shadow_lessons',
        'attraction': 'attraction',
        'lesson': 'lesson',
        'wound': 'wound',
        'partner': 'partner',
        'karmic': 'karmic',
        'past life': 'past_life',
        'past_life': 'past_life',
        'resolution': 'resolution',
        'money flow': 'money_flow',
        'money_flow': 'money_flow',
        'block': 'block',
        'activation': 'activation',
        'life purpose': 'life_purpose',
        'life_purpose': 'life_purpose',
        'gifts': 'gifts',
        'mission': 'mission',
        'current cycle': 'current_cycle',
        'current_cycle': 'current_cycle',
        'next phase': 'next_phase',
        'next_phase': 'next_phase',
        'yearly': 'yearly',
    }
    
    if section_part in section_map:
        section = section_map[section_part]
    else:
        # Fallback to cleaning section names dynamically (spaces to underscores)
        section = section_part.replace(' ', '_').strip('_')
        if not section:
            section = 'meaning'
            
    return pos, module, section


def main():
    print("=" * 60)
    print("      Soul Blueprint Matrix - Excel Database Sync Tool")
    print("=" * 60)
    
    excel_file = get_excel_file()
    if not excel_file:
        print("[ERROR] Could not find any interpretations spreadsheet!")
        print("Please place your Excel file (e.g. interpretations.xlsx) in this folder.")
        input("\nPress Enter to exit...")
        sys.exit(1)
        
    print(f"[INFO] Found Excel database: {excel_file}")
    
    # Check if we can open the Excel file (prevents crashes from Excel file locks)
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    except PermissionError:
        print("\n[ERROR] Permission denied opening the Excel file.")
        print("This usually happens if the file is currently open in Microsoft Excel.")
        print("Please CLOSE the Excel file and run this tool again.")
        input("\nPress Enter to exit...")
        sys.exit(1)
    except Exception as e:
        print(f"\n[ERROR] Failed to load the spreadsheet: {e}")
        input("\nPress Enter to exit...")
        sys.exit(1)
        
    # Read the first sheet or the Master_Database sheet
    sheet_name = "Master_Database" if "Master_Database" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    print(f"[INFO] Reading sheet: '{sheet_name}'")
    
    # Read rows
    try:
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        print(f"[ERROR] Failed to read cell contents: {e}")
        wb.close()
        input("\nPress Enter to exit...")
        sys.exit(1)
    finally:
        wb.close()
        
    if not rows:
        print("[ERROR] The spreadsheet is completely empty!")
        input("\nPress Enter to exit...")
        sys.exit(1)
        
    # Parse headers dynamically
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    
    required_cols = {
        'position': ['position', 'pos', 'node'],
        'position meaning': ['position meaning', 'meaning', 'description'],
        'section': ['section', 'category', 'tab'],
        'number': ['number', 'val', 'num', 'key'],
        'interpretation text': ['interpretation text', 'text', 'content', 'interpretation']
    }
    
    col_indices = {}
    for col_key, aliases in required_cols.items():
        found_idx = None
        for alias in aliases:
            if alias in headers:
                found_idx = headers.index(alias)
                break
        if found_idx is None:
            # Try partial matching if exact matches failed
            for i, h in enumerate(headers):
                if any(alias in h for alias in aliases):
                    found_idx = i
                    break
        if found_idx is None:
            print(f"[ERROR] Could not find the '{col_key}' column in Excel!")
            print(f"Spreadsheet columns must include headers like: Position, Section, Number, Interpretation Text.")
            print(f"Current columns found: {headers}")
            input("\nPress Enter to exit...")
            sys.exit(1)
        col_indices[col_key] = found_idx
        
    pos_idx = col_indices['position']
    meaning_idx = col_indices['position meaning']
    sec_idx = col_indices['section']
    num_idx = col_indices['number']
    text_idx = col_indices['interpretation text']
    
    # 2. Back up the old CSV before editing
    csv_filename = "interpretations.csv"
    backup_path = create_backup(csv_filename)
    if backup_path:
        print(f"[OK] Pre-existing CSV backed up to: {backup_path}")
        
    # 3. Export to CSV
    mapped_count = 0
    empty_count = 0
    invalid_rows = []
    
    # Track positions and sections for output summary
    active_positions = set()
    
    try:
        with open(csv_filename, 'w', encoding='utf-8', newline='') as csvfile:
            writer = csv.writer(csvfile)
            # Write standardized headers
            writer.writerow(['position', 'module', 'section', 'number', 'text'])
            
            for row_num, row in enumerate(rows[1:], start=2):
                p_val = row[pos_idx] if pos_idx < len(row) else None
                p_mean = row[meaning_idx] if meaning_idx < len(row) else None
                sec_val = row[sec_idx] if sec_idx < len(row) else None
                num_val = row[num_idx] if num_idx < len(row) else None
                text_val = row[text_idx] if text_idx < len(row) else None
                
                # Skip rows that have completely empty position
                if p_val is None or str(p_val).strip() == "":
                    continue
                # If text is empty, count it as skipped empty row
                if text_val is None or str(text_val).strip() == "":
                    empty_count += 1
                    continue
                    
                # Safe conversions (bypass integer conversion for 3-number programs)
                p_val_str = str(p_val).strip() if p_val is not None else ""
                is_program = '-' in p_val_str or ',' in p_val_str or p_val_str.upper() == 'PROGRAM'
                
                if is_program:
                    num = str(num_val).strip() if num_val is not None else ""
                else:
                    try:
                        num = int(float(num_val))
                    except (ValueError, TypeError):
                        invalid_rows.append((row_num, row))
                        continue
                    
                # Map columns
                pos, module, section = map_excel_to_csv_columns(p_val, p_mean, sec_val)
                text_clean = str(text_val).strip()
                
                writer.writerow([pos, module, section, num, text_clean])
                mapped_count += 1
                active_positions.add(pos)
                
    except PermissionError:
        print(f"\n[ERROR] Permission denied: Could not write to '{csv_filename}'.")
        print("Please close any applications (like text editors or Excel) that might have the file locked.")
        input("\nPress Enter to exit...")
        sys.exit(1)
    except Exception as e:
        print(f"\n[ERROR] Failed to write database: {e}")
        input("\nPress Enter to exit...")
        sys.exit(1)
        
    print("\n" + "=" * 60)
    print("                      SUCCESSFUL SYNC")
    print("=" * 60)
    print(f"[OK] Database updated: {csv_filename}")
    print(f"[OK] Total rows imported: {mapped_count}")
    print(f"[OK] Empty rows skipped: {empty_count}")
    
    if active_positions:
        print(f"[OK] Positions synced: {', '.join(sorted(list(active_positions)))}")
        
    if invalid_rows:
        print(f"\n[WARNING] Skipped {len(invalid_rows)} rows due to invalid 'Number' values:")
        for r_num, row in invalid_rows[:5]:
            print(f"  - Row {r_num}: Number='{row[num_idx]}' (Position {row[pos_idx]})")
        if len(invalid_rows) > 5:
            print(f"  - ... and {len(invalid_rows) - 5} more rows.")
            
    print("\nWorkflow ready. You can now reload your Soul Blueprint Matrix in Chrome.")
    print("=" * 60)
    
    # We add a pause when running in batch mode, but don't hold the shell if run directly
    if len(sys.argv) > 1 and sys.argv[1] == '--batch':
        input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
